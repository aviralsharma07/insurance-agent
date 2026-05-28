"""IRDAI Historical Performance Data Scraper

Extracts insurer-wise performance data from IRDAI Handbook of Insurance Statistics
across all available years (2007-08 to 2024-25) from 3 format eras:
  - XLSX (2020-21 to 2024-25): structured spreadsheets, openpyxl
  - XLS  (2009-10, 2017-18): old Excel, xlrd
  - PDF  (remaining years): text-based tables, pdfplumber

Output: data/irdai_historical.json
"""

import json
import re
import os
import sys
import zipfile
import tempfile
import shutil
import logging
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Optional, Any

import pdfplumber
import openpyxl
import xlrd

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger("irdai_scraper")

# ── Paths ────────────────────────────────────────────────────────────────────
HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent.parent
DATA_DIR = PROJECT_ROOT / "data"
OUTPUT_PATH = DATA_DIR / "irdai_historical.json"
TPA_OUTPUT_PATH = DATA_DIR / "irdai_historical_tpa_network.json"

# Source data (outside project root)
POLICY_DATA = Path("/Users/aviralsharma/Personal Projects/policy_data")
IRDAI_DIR = POLICY_DATA / "_irdai_reference"
HANDBOOK_DIR = IRDAI_DIR / "data"
REPORT_DIR = IRDAI_DIR / "reports"

# ── Insurer Name Normalization ───────────────────────────────────────────────
# Maps raw handbook names to canonical short names (matching UIN scraper style)
SUFFIX_PATTERNS = [
    r"\s+Co\.?\s*Ltd\.?$",
    r"\s+C\.?\s*Ltd\.?$",
    r"\s+Ltd\.?$",
    r"\s+Limited$",
    r"\s+Insurance\s+Co\.?\s*Ltd\.?$",
    r"\s+Insurance\s+C\.?\s*Ltd\.?$",
    r"\s+General\s+Insurance\s+Co\.?\s*Ltd\.?$",
    r"\s+General\s+Insurance\s+C\.?\s*Ltd\.?$",
    r"\s+Health\s+Insurance\s+Co\.?\s*Ltd\.?$",
    r"\s+Health\s+Insurance\s+C\.?\s*Ltd\.?$",
]

CANONICAL_NAMES = {
    "national insurance": "National Insurance",
    "new india assurance": "New India Assurance",
    "oriental insurance": "Oriental Insurance",
    "united india insurance": "United India Insurance",
    "hdfc ergo": "HDFC ERGO",
    "hdfc ergo general insurance": "HDFC ERGO",
    "icici lombard": "ICICI Lombard",
    "icici lombard general insurance": "ICICI Lombard",
    "star health": "Star Health",
    "star health and allied insurance": "Star Health",
    "niva bupa": "Niva Bupa",
    "niva bupa health insurance": "Niva Bupa",
    "care health": "Care Health",
    "care health insurance": "Care Health",
    "aditya birla health insurance": "Aditya Birla Health Insurance",
    "bajaj allianz": "Bajaj Allianz General Insurance",
    "bajaj allianz general insurance": "Bajaj Allianz General Insurance",
    "tata aig": "Tata AIG General Insurance",
    "tata aig general insurance": "Tata AIG General Insurance",
    "iffco tokio": "IFFCO Tokio General Insurance",
    "iffco-tokio": "IFFCO Tokio General Insurance",
    "iffco tokio general insurance": "IFFCO Tokio General Insurance",
    "reliance general insurance": "Reliance General Insurance",
    "royal sundaram": "Royal Sundaram General Insurance",
    "royal sundaram general insurance": "Royal Sundaram General Insurance",
    "chola": "Cholamandalam MS General Insurance",
    "cholamandalam": "Cholamandalam MS General Insurance",
    "cholamandalam ms general insurance": "Cholamandalam MS General Insurance",
    "future generali": "Future Generali India Insurance",
    "future generali india insurance": "Future Generali India Insurance",
    "bharti axa": "Bharti AXA General Insurance",
    "bharti axa general insurance": "Bharti AXA General Insurance",
    "sbi general insurance": "SBI General Insurance",
    "go digit": "Go Digit General Insurance",
    "go digit general insurance": "Go Digit General Insurance",
    "acko general insurance": "Acko General Insurance",
    "liberty general insurance": "Liberty General Insurance",
    "liberty videocon": "Liberty General Insurance",
    "universal sompo": "Universal Sompo General Insurance",
    "universal sompo general insurance": "Universal Sompo General Insurance",
    "kotak mahindra general insurance": "Kotak Mahindra General Insurance",
    "kotak general insurance": "Kotak Mahindra General Insurance",
    "magma hdi": "Magma HDI General Insurance",
    "magma hdi general insurance": "Magma HDI General Insurance",
    "shriram": "Shriram General Insurance",
    "shri ram general insurance": "Shriram General Insurance",
    "raheja qbe": "Raheja QBE General Insurance",
    "raheja qbe general insurance": "Raheja QBE General Insurance",
    "navi general insurance": "Navi General Insurance",
    "edelweiss general insurance": "Edelweiss General Insurance",
    "dhfl general insurance": "DHFL General Insurance",
    "apollo munich health insurance": "Apollo Munich Health Insurance",
    "cigna ttk health insurance": "CIGNA TTK Health Insurance",
    "manipalcigna health insurance": "ManipalCigna Health Insurance",
    "kshema general insurance": "Kshema General Insurance",
    "kshemageneral insurance": "Kshema General Insurance",
    "ecgc": "ECGC",
    "ecgc ltd.": "ECGC",
    "agriculture insurance company of india": "Agriculture Insurance Co. of India Ltd.",
    "agriculture insurance": "Agriculture Insurance Co. of India Ltd.",
    "new india": "New India Assurance",
    "star health & allied insurance": "Star Health",
    "star health and allied insurance": "Star Health",
    "zurich kotak general insurance co.(india)": "Zurich Kotak General Insurance Co. (India) Ltd.",
    "zurich kotak general insurance co. (india)": "Zurich Kotak General Insurance Co. (India) Ltd.",
    "general insurance corporation of india": "General Insurance Corporation of India",
    "hdfc ergo health insurance": "HDFC ERGO Health Insurance",
    "lt general insurance": "L&T General Insurance",
    "l&t general insurance": "L&T General Insurance",
    "max bupa health insurance": "Max Bupa Health Insurance",
    "narayana health insurance": "Narayana Health Insurance",
    "galaxy health insurance": "Galaxy Health Insurance",
    "galaxy health and allied insurance company": "Galaxy Health Insurance",
    "reliance health insurance": "Reliance Health Insurance",
    "magma general insurance": "Magma General Insurance",
    "stand-alone health insurers": "Stand-alone Health Insurers",
    "standalone health insurers": "Stand-alone Health Insurers",
}


def normalize_insurer(name: str) -> str:
    name = str(name or "").strip()
    if not name:
        return "Unknown"
    if is_footnote_or_header(name):
        return "Unknown"
    key = name.lower().strip()
    # Strip trailing footnote markers (#, $, %, &, *, ^, @)
    key = re.sub(r"[\#$%&\*\^@]+$", "", key).strip()
    # Strip leading footnote markers
    key = re.sub(r"^[\#$%&\*\^@]+", "", key).strip()
    for p in SUFFIX_PATTERNS:
        key = re.sub(p, "", key, flags=re.IGNORECASE).strip()
    key = re.sub(r"^the\s+", "", key).strip()
    key = re.sub(r"\s+", " ", key).strip()
    if key in CANONICAL_NAMES:
        return CANONICAL_NAMES[key]
    # Return original name (not key) with markers and suffixes stripped
    clean = name.strip()
    clean = re.sub(r"[\#$%&\*\^@~]+$", "", clean).strip()
    for p in SUFFIX_PATTERNS:
        clean = re.sub(p, "", clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r"\s+", " ", clean).strip()
    return clean or name.strip()


PURE_NON_LIFE = {
    "New India Assurance",
    "Oriental Insurance",
    "United India Insurance",
    "Bajaj Allianz General Insurance",
    "Tata AIG General Insurance",
    "IFFCO Tokio General Insurance",
    "Reliance General Insurance",
    "Royal Sundaram General Insurance",
    "Cholamandalam MS General Insurance",
    "Future Generali India Insurance",
    "Bharti AXA General Insurance",
    "SBI General Insurance",
    "Go Digit General Insurance",
    "Acko General Insurance",
    "Liberty General Insurance",
    "Universal Sompo General Insurance",
    "Kotak Mahindra General Insurance",
    "Magma HDI General Insurance",
    "Shriram General Insurance",
    "Raheja QBE General Insurance",
    "Navi General Insurance",
    "Edelweiss General Insurance",
    "DHFL General Insurance",
    "Kshema General Insurance",
}

STANDALONE_HEALTH = {
    "Star Health",
    "Niva Bupa",
    "Care Health",
    "Aditya Birla Health Insurance",
    "Apollo Munich Health Insurance",
    "CIGNA TTK Health Insurance",
    "ManipalCigna Health Insurance",
}

PSU_INSURERS = {
    "National Insurance",
    "New India Assurance",
    "Oriental Insurance",
    "United India Insurance",
}

ALL_CANONICAL = sorted(set(CANONICAL_NAMES.values()) | {"National Insurance"})


# ── Helper: parse numeric value ──────────────────────────────────────────────
def parse_num(val) -> Optional[float]:
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace(" ", "")
    if not s or s in ("-", "—", "–", "NA", "N/A", ""):
        return None
    s = s.replace("(", "-").replace(")", "")
    try:
        return float(s)
    except ValueError:
        return None


def is_total_row(insurer_name: str) -> bool:
    name_lower = insurer_name.lower().strip()
    if any(
        kw in name_lower
        for kw in [
            "total",
            "sub-total",
            "subtotal",
            "public sector",
            "private sector",
            "grand total",
            "specialized",
            "private insurers",
            "public insurers",
        ]
    ):
        return True
    return False


def is_footnote_or_header(name: str) -> bool:
    """Check if a name is a footnote, table header, or non-insurer text."""
    if not name or not name.strip():
        return True
    s = name.strip()
    # Starts with special footnote markers (including quotes)
    if re.match(r"""^[#$%&\*\'\(\)\+\-\/^"]""", s):
        return True
    # Starts with a digit followed by text (PDF table artifact like "75 Health Insurance")
    if re.match(r"^\d+\s+[A-Za-z]", s):
        return True
    # Starts with a number (likely a serial number or footnote number)
    if re.match(r"^\d+\.?\s*$", s):
        return True
    # Table titles
    if re.match(r"^TABLE\s+\d+", s, re.IGNORECASE):
        return True
    if re.match(r"^STATUS OF", s, re.IGNORECASE):
        return True
    if re.match(r"^S\.?NO\.?", s, re.IGNORECASE):
        return True
    if re.match(r"^PART\s", s, re.IGNORECASE):
        return True
    if re.match(r"^PROFILE$", s, re.IGNORECASE):
        return True
    if re.match(r"^SPECIALI", s, re.IGNORECASE):
        return True
    # Column header names
    if s.lower() in (
        "insurer",
        "insurers",
        "particulars",
        "category",
        "type of insurer",
        "name of the insurer",
        "reinsurer",
        "reinsurers",
    ):
        return True
    # Contains footnote indicators
    if re.search(
        r"(?:NA\s+indicates|not\s+in\s+operation|renamed\s+as|erstwhile|formerly\s+known\s+as|merged\s+with|w\.e\.f\.|started\s+operations|demerger\s+of|carried\s+forward|brought\s+forward|\"-\"\s+indicates|indicates\s+the\s+company)",
        s,
        re.IGNORECASE,
    ):
        return True
    # Contains age/period footnote patterns
    if re.search(
        r"(?:age\s+is\s+\d|years?\s+to\s+[<\d]|years?\*{1,2})", s, re.IGNORECASE
    ):
        return True
    # Pure numbers or "N/A" only
    if re.match(r"^[\d\s,.()%-]+$", s):
        return True
    # Financial line items (PDF extraction artifacts)
    if re.search(
        r"(?:FY\s+[`(]|Per\s+Cent|Commission\s+expenses|Assets\s+Under\s+Management|Operating\s+(?:expenses|profit)|Income\s+from\s+investments|Underwriting\s+Profit|Equity\s+Share\s+Capital|Market\s+share\s+of\s+PSUs|Net\s+Retentions|Segment[- ]wise|BUSINESS\s+FIGURES|Figure\s+[ivxlcdm]|Sl\.?\s*Page|No\.?\s*No\.?$|REINSURERS)",
        s,
        re.IGNORECASE,
    ):
        return True
    # Standalone letter + word pattern (PDF truncation artifacts like "ADITYA B", "BAJAJ AL")
    if re.match(r"^[A-Z]{3,10}\s+[A-Z]{1,3}$", s):
        return True
    # Category separator — any all-caps short text that isn't a known insurer
    if s.upper() == s and len(s) < 30 and len(s.split()) <= 3:
        known_sections = {
            "public",
            "private",
            "public sector",
            "private sector",
            "public sector insurers",
            "private sector insurers",
            "standalone health",
            "specialized",
            "others",
        }
        if s.lower() in known_sections:
            return True
        # Uppercase with 2-3 words that are NOT known insurers — likely a PDF artifact
        words = s.split()
        insurer_keywords = {
            "insurance",
            "general",
            "health",
            "ltd",
            "co",
            "company",
            "india",
            "corporation",
            "limited",
            "reinsurance",
        }
        has_insurer_kw = any(w.lower() in insurer_keywords for w in words)
        if not has_insurer_kw and len(words) >= 2:
            return True
    # Truncated PDF text — single uppercase letter at end (e.g. "Apollo D", "Apollo M")
    if re.match(r"^[A-Z][a-z]+[\s']+[A-Z][a-z]*[\s']*[A-Z]?$", s) and len(s) < 15:
        words = s.split()
        if len(words) >= 2 and len(words[-1]) <= 2:
            return True
    # "Engineering FY In Per Cen" — financial line items from PDF
    if re.search(
        r"(?:FY\s+[`(In]|Per\s+Cen|Figure\s+[ivxlcdm]|Sl\.?\s*Page|Prof\.\s+No\.)",
        s,
        re.IGNORECASE,
    ):
        return True
    # "S. No." or "Sl. No." with space
    if re.match(r"^(S|Sl)\.?\s*No\.?$", s, re.IGNORECASE):
        return True
    if re.match(r"^No\.?\s+of\s*$", s, re.IGNORECASE):
        return True
    # Starts with "Note:", "NA:", or "~" (word boundary to avoid matching "National")
    if re.match(r"^(Note\b|NA\b|N\.A\.)\s*[:.]?\s*", s, re.IGNORECASE):
        return True
    if s.startswith("~"):
        return True
    # Starts with em-dash, en-dash, or similar
    if s and ord(s[0]) in (0x2013, 0x2014, 0x2015):
        return True
    # Single character or single letter
    if len(s) <= 1:
        return True
    # "No. of" — table page descriptors
    if re.match(r"^No\.?\s+of\s+", s, re.IGNORECASE):
        return True
    # Financial/administrative line items (not insurer names)
    if re.search(
        r"(?:Profit.*Loss|branch\s+offices|registered\s+broch|channel\s+wise|state[- ]?wise)",
        s,
        re.IGNORECASE,
    ):
        return True
    # Table headers / descriptors
    if re.match(
        r"^(Gross Direct Premium|Incurred Claims Ratio|Net Retentions|Segment[- ]Wise)",
        s,
        re.IGNORECASE,
    ):
        return True
    # "Marine Cargo FY", "Aviation FY", etc. — segment descriptors
    if re.match(
        r"^(Fire|Marine|Motor|Health|Aviation|Engineering|Others)\s+FY",
        s,
        re.IGNORECASE,
    ):
        return True
    # Table numbers with letter suffixes: "86A", "87A", "88A"
    if re.match(r"^\d{2,3}[A-Z]\s+", s):
        return True
    # PDF truncated names: last word is 2-3 chars and likely incomplete
    words = s.strip().split()
    if len(words) >= 2:
        last = words[-1].strip(".")  # strip trailing period for comparison
        common_short = {
            "ltd",
            "co",
            "inc",
            "axa",
            "sbi",
            "ttk",
            "qbe",
            "hdi",
            "ms",
            "aig",
        }
        # If last word is 2-3 chars and not a known short word, likely truncated
        if 2 <= len(last) <= 3 and last.lower() not in common_short:
            return True
        # If last word is 2-4 chars and the full name is not in canonical_values
        if 2 <= len(last) <= 4 and last.lower() not in common_short:
            cv = {v.lower() for v in CANONICAL_NAMES.values()}
            if s.strip().lower() not in cv:
                return True
    return False


# ══════════════════════════════════════════════════════════════════════════════
#  TABLE REGISTRY — maps (handbook_year, field) → table identifiers
# ══════════════════════════════════════════════════════════════════════════════

# XLSX era (2020-21 onwards): Part II sheet names differ by year
XLSX_PART_II_TABLES = {
    2025: {
        "gwp": "40",
        "icr_all": "44",
        "solvency": "49",
        "claims": "53",
        "grievances": " 56",
    },
    2024: {
        "gwp": "40",
        "icr_all": "44",
        "solvency": "49",
        "claims": "53",
        "grievances": " 56",
    },
    2023: {
        "gwp": "42",
        "icr_all": "46",
        "solvency": "51",
        "claims": "55",
        "grievances": "58",
    },
    2022: {
        "gwp": "44",
        "icr_all": "48",
        "solvency": "53",
        "claims": "57",
        "grievances": "60",
    },
    2021: {
        "gwp": "50",
        "icr_all": "55",
        "solvency": "59",
        "claims": None,
        "grievances": "84",
        "health_persons": "63",
        "health_icr": "64",
    },
}

XLSX_PART_III_TABLES = {
    2025: {
        "health_persons": "58",
        "health_icr": "62",
        "tpa_network": "77",
        "network_providers": "78",
    },
    2024: {
        "health_persons": "58",
        "health_icr": "62",
        "tpa_network": "77",
        "network_providers": "78",
    },
    2023: {
        "health_persons": "60",
        "health_icr": "64",
        "tpa_network": "77",
        "network_providers": "78",
    },
    2022: {
        "health_persons": "62",
        "health_icr": "66",
        "tpa_network": "77",
        "network_providers": "78",
    },
    2021: {
        "health_persons": None,
        "health_icr": None,
        "tpa_network": None,
        "network_providers": None,
    },
}

# XLS era (2017-18): filename fragments
XLS_2018_TABLES = {
    "gwp": "Gross_Direct_Premium",
    "icr": "Incurred_Claims_Ratio",
    "solvency": "Solvency_Ratio_of_General_Insurers",
    "health_persons": "Health Insurance_Excl_Travel_ PA_Gross Premium_number of persons covered",
    "health_icr": "HI_Excl_Trvl_PA_Incurred_Claims_Ratio",
    "grievances": "Status_of_Grievances_General_Insurers",
}

XLS_2010_TABLES = {
    "gwp": "Gross_Direct_Premium_of_Non_Life_Insurance",
    "icr": "Incurred_Claims_Ratio",
    "solvency": "solvency ratios of non-life insurers",
    "grievances": "Status of grievances -Non-Life Insurers",
}

# PDF era: table numbers vary; we parse TOC to find them dynamically
# Known mappings for 2018-19 and 2019-20 (same numbering as 2017-18 XLS)
PDF_TABLE_KWARGS = {
    "gwp": [
        "gross direct premium",
        "within & outside india",
        "gross direct premium of non-life",
        # Reversed-text search (right-to-left PDF encoding)
        "muimerp tcerid ssorg",
        "aidni edistuo & nihtiw",
    ],
    "icr": ["incurred claims ratio"],
    "solvency": ["solvency ratio of general"],
    "health_persons": ["health insurance.*gross premium.*number of persons"],
    "health_icr": ["health insurance.*net earned premium.*incurred claims"],
    "grievances": ["status of grievances"],
    "claims": ["status of claims"],
}


# ── Reversed Text Detection & Fixing ─────────────────────────────────────────
# Some PDF handbooks (2010-11 through 2013-14, 2018-19/2019-20 GWP pages)
# use right-to-left PDF encoding. pdfplumber extracts each word character-reversed.
# Example: "TABLE 49: GROSS DIRECT PREMIUM INCOME" -> "ELBAT :94 SSORG TCERID MUIMERP EMOCNI"
# We detect this by looking for reversed year patterns YY-YYYY instead of YYYY-YY.


def _is_reversed_text(text: str) -> bool:
    """Detect if page text is reversed (right-to-left PDF encoding)."""
    rev = len(re.findall(r"\b(\d{2}-\d{4})\b", text[:2000]))
    norm = len(re.findall(r"\b(\d{4}-\d{2})\b", text[:2000]))
    return rev > norm


def _fix_reversed_text(text: str) -> str:
    """Fix reversed text by reversing each word character-by-character."""
    lines = text.split("\n")
    result = []
    for line in lines:
        words = line.strip().split()
        fixed = []
        for w in words:
            rw = w[::-1]
            # Fix reversed years YY-YYYY back to YYYY-YY
            rw = re.sub(r"\b(\d{2})-(\d{4})\b", r"\2-\1", rw)
            fixed.append(rw)
        result.append(" ".join(fixed))
    return "\n".join(result)


def _extract_landscape_table(page, handbook_year: int) -> dict:
    """Extract data from landscape-format table (years as rows, insurers as columns)."""
    words = page.extract_words(keep_blank_chars=True, x_tolerance=3)
    if not words:
        return {}

    sample_text = " ".join(w["text"] for w in words[: max(50, min(200, len(words)))])
    is_reversed = _is_reversed_text(sample_text)
    if is_reversed:
        for w in words:
            rw = w["text"][::-1]
            rw = re.sub(r"\b(\d{2})-(\d{4})\b", r"\2-\1", rw)
            w["text"] = rw

    Y_TOLERANCE = 7
    raw_lines = defaultdict(list)
    for w in words:
        y_key = round(w["top"] / Y_TOLERANCE) * Y_TOLERANCE
        raw_lines[y_key].append(w)

    sorted_lines = []
    for y in sorted(raw_lines.keys()):
        line_words = sorted(raw_lines[y], key=lambda w: w["x0"])
        sorted_lines.append((y, line_words))

    # Identify data rows + collect all value positions per year column
    data_rows = {}
    for y, line_words in sorted_lines:
        first_word = line_words[0]["text"] if line_words else ""
        if re.match(r"^\d{4}-\d{2}$", first_word):
            year = first_word
            values = [(w["x0"], w["text"]) for w in line_words[1:]]
            if values:
                data_rows[year] = values

    if not data_rows:
        return {}

    # Build column clusters by grouping x0 positions that are close together
    # For each year, collect all value x0s — then cluster across years
    all_x0s = set()
    for year, vals in data_rows.items():
        for x0, _ in vals:
            all_x0s.add(x0)

    sorted_x0s = sorted(all_x0s)
    if not sorted_x0s:
        return {}

    COL_GAP = 8  # max gap between consecutive x0s in same column cluster
    col_clusters = []  # list of (x0_center, x0_start, x0_end) for each column
    cluster_start = sorted_x0s[0]
    cluster_vals = [sorted_x0s[0]]
    for x0 in sorted_x0s[1:]:
        if x0 - cluster_vals[-1] > COL_GAP:
            center = (cluster_start + cluster_vals[-1]) / 2
            x0_end = (cluster_start + x0) / 2
            col_clusters.append((center, cluster_start, x0_end))
            cluster_start = x0
            cluster_vals = [x0]
        else:
            cluster_vals.append(x0)
    center = (cluster_start + cluster_vals[-1]) / 2
    col_clusters.append((center, cluster_start, sorted_x0s[-1] + 1))

    # Find insurer names by scanning bottom portion of page
    # Look for words inside each column's x0 range
    last_data_y = max(
        y
        for y, _ in sorted_lines
        if any(re.match(r"^\d{4}-\d{2}$", w["text"]) for w in _)
    )

    insurer_by_x0 = {}  # x0_center -> canonical name
    for y, line_words in sorted_lines:
        if y <= last_data_y:
            continue
        for center, x0_start, x0_end in col_clusters:
            if center in insurer_by_x0:
                continue
            col_words = [
                w["text"]
                for w in line_words
                if w["x0"] >= x0_start and w["x0"] < x0_end
            ]
            col_text = " ".join(col_words).strip()
            if len(col_text) < 2:
                continue
            canonical = normalize_insurer(col_text)
            if canonical != "Unknown" and canonical not in insurer_by_x0.values():
                insurer_by_x0[center] = canonical

    # If enough insurers found, map data to them
    if len(insurer_by_x0) >= 3:
        sorted_centers = sorted(insurer_by_x0.keys())
        result = defaultdict(dict)
        for year, vals in data_rows.items():
            for i, (x0, v) in enumerate(vals):
                val = parse_num(v)
                if val is not None and i < len(sorted_centers):
                    ins = insurer_by_x0[sorted_centers[i]]
                    result[ins][year] = val
        return dict(result)

    return {}


def _rebuild_page_text(page) -> str:
    """Rebuild text lines from word-level extraction for reversed-text pages.
    Groups words by y-position and sorts by x0 to reconstruct actual reading order lines."""
    words = page.extract_words(keep_blank_chars=True, x_tolerance=3)
    if not words:
        return ""

    sample = " ".join(w["text"] for w in words[:50])
    if _is_reversed_text(sample):
        for w in words:
            rw = w["text"][::-1]
            rw = re.sub(r"\b(\d{2})-(\d{4})\b", r"\2-\1", rw)
            w["text"] = rw

    Y_TOL = 5
    by_y = defaultdict(list)
    for w in words:
        y_key = round(w["top"] / Y_TOL) * Y_TOL
        by_y[y_key].append(w)

    text_lines = []
    for y in sorted(by_y.keys()):
        line_words = sorted(by_y[y], key=lambda w: w["x0"])
        text_lines.append(" ".join(w["text"] for w in line_words))
    return "\n".join(text_lines)


def pdf_extract_gwp_table(page, handbook_year: int) -> dict:
    """Extract GWP from a PDF table page, handling both normal and landscape formats."""
    text = page.extract_text() or ""

    if _is_reversed_text(text):
        landscape = _extract_landscape_table(page, handbook_year)
        if landscape:
            return landscape
        text = _rebuild_page_text(page)

    lines = text.strip().split("\n")
    result = {}

    header_idx = None
    year_cols_map = {}

    for li, line in enumerate(lines):
        years_found = re.findall(r"\b(\d{4}-\d{2})\b", line)
        if len(years_found) >= 3:
            header_idx = li
            for y in years_found:
                pos = line.find(y)
                year_cols_map[pos] = y
            break

    if not year_cols_map:
        for li, line in enumerate(lines):
            years_found = re.findall(r"\b(\d{4})\b", line)
            valid = [y for y in years_found if 2000 <= int(y) <= 2030]
            if len(valid) >= 3:
                for y in valid:
                    pos = line.find(y)
                    yr = int(y)
                    year_cols_map[pos] = f"{yr - 1}-{str(yr)[2:]}"
                break

    if not year_cols_map:
        return result

    sorted_positions = sorted(year_cols_map.keys())

    for li in range((header_idx or 0) + 1, len(lines)):
        raw = lines[li]
        if not raw.strip():
            continue
        if is_total_row(raw[:60]):
            continue
        if re.match(r"^[\d\s.,()%-]+$", raw.strip()):
            continue

        ins_name = (
            raw[: sorted_positions[0]].strip() if sorted_positions else raw.strip()
        )
        if not ins_name:
            continue
        try:
            float(ins_name.replace(",", "").replace("-", "0"))
            continue
        except ValueError:
            pass

        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            words = ins_name.split()[:3]
            if words:
                canonical = normalize_insurer(" ".join(words))
        if canonical == "Unknown":
            continue

        if canonical not in result:
            result[canonical] = {}

        for i, pos in enumerate(sorted_positions):
            fy = year_cols_map[pos]
            start = pos
            end = sorted_positions[i + 1] if i + 1 < len(sorted_positions) else len(raw)
            val_str = raw[start:end].strip()
            val = parse_num(val_str)
            if val is not None:
                result[canonical][fy] = val

    return result


def pdf_extract_icr_table(page, handbook_year: int) -> dict:
    """Extract ICR values from PDF table pages, handling both normal and reversed text.

    Supports two table formats:
      - Landscape (years as rows): handled by _extract_landscape_table, NOT this function
      - Portrait (insurer as rows, years as columns): two sub-types:
        (a) Years embedded in text with positions > 0: character-position extraction
        (b) Years at start of line (position 0): whitespace-based extraction
    """
    text = page.extract_text() or ""

    if _is_reversed_text(text):
        text = _rebuild_page_text(page)

    lines = text.strip().split("\n")
    result = {}

    # Find year header line and count ALL year columns (including duplicates for NEP+ICR)
    year_header_line = None
    total_year_count = 0
    for line in lines:
        years_found = re.findall(r"\b(\d{4}-\d{2})\b", line)
        if len(years_found) >= 2:
            year_header_line = line
            total_year_count = len(years_found)
            break

    if not year_header_line:
        for line in lines:
            years_found = re.findall(r"\b(\d{4})\b", line)
            valid = [y for y in years_found if 2000 <= int(y) <= 2030]
            if len(valid) >= 2:
                year_header_line = line
                total_year_count = len(valid)
                break

    if not year_header_line or total_year_count < 2:
        return result

    # Get unique year values in order (first half = NEP years, second half = ICR years)
    all_years = re.findall(r"\b(\d{4}-\d{2})\b", year_header_line)
    if not all_years:
        all_year_nums = re.findall(r"\b(\d{4})\b", year_header_line)
        valid = [y for y in all_year_nums if 2000 <= int(y) <= 2030]
        all_years = [f"{int(y) - 1}-{str(y)[2:]}" for y in valid]
    if not all_years or len(all_years) < 2:
        return result

    n_total = len(all_years)
    has_duplicates = len(all_years) != len(set(all_years))
    # If duplicates exist (e.g. "2013-14 2014-15 2015-16 2013-14 2014-15 2015-16"):
    #   NEP+ICR paired columns — take last half as ICR values
    # If no duplicates (e.g. "2013-14...2019-20"): single column set — all values are ICR
    n_icr_cols = n_total // 2 if has_duplicates else n_total
    icr_years = all_years[: min(n_icr_cols, len(all_years))]

    # Check if years start at/near position 0 (portrait table with years at line start)
    first_match = re.search(r"\b(\d{4}-\d{2})\b", year_header_line)
    is_portrait = first_match is not None and first_match.start() < 5

    if not is_portrait:
        # ── Character-position based extraction ──
        year_cols_map = {}
        for m in re.finditer(r"\b(\d{4}-\d{2})\b", year_header_line):
            pos = m.start()
            if pos not in year_cols_map:
                year_cols_map[pos] = m.group(1)
        if not year_cols_map:
            for m in re.finditer(r"\b(\d{4})\b", year_header_line):
                pos = m.start()
                yr = int(m.group(1))
                fy_key = f"{yr - 1}-{str(yr)[2:]}"
                if pos not in year_cols_map:
                    year_cols_map[pos] = fy_key

        sorted_positions = sorted(year_cols_map.keys())
        if not sorted_positions:
            return result

        for line in lines:
            raw = line.strip()
            if not raw or is_total_row(raw[:60]):
                continue
            if re.match(r"^[\d\s.,()%-]+$", raw):
                continue

            ins_name = raw[: sorted_positions[0]].strip()
            if not ins_name:
                continue
            try:
                float(ins_name.replace(",", "").replace("-", "0"))
                continue
            except ValueError:
                pass

            canonical = normalize_insurer(ins_name)
            if canonical == "Unknown":
                continue
            if canonical not in result:
                result[canonical] = {}

            for i, pos in enumerate(sorted_positions):
                fy = year_cols_map[pos]
                start = pos
                end = (
                    sorted_positions[i + 1]
                    if i + 1 < len(sorted_positions)
                    else len(raw)
                )
                val_str = raw[start:end].strip()
                val = parse_num(val_str)
                if val is not None:
                    result[canonical][fy] = val
    else:
        # ── Whitespace-based extraction for portrait tables ──
        # Table structure: [insurer words] [NEP val1] [NEP val2] ... [ICR val1] [ICR val2] ...
        # Number of values per insurer = total_year_count (NEP + ICR)
        # ICR values are the last n_unique values

        for line in lines:
            raw = line.strip()
            if not raw or is_total_row(raw[:60]):
                continue
            if re.match(r"^[\d\s.,()%-]+$", raw):
                continue

            tokens = raw.split()
            if len(tokens) < n_icr_cols + 2:  # At least insurer name + 1 value
                continue

            # Find the first numeric token — everything before it is the insurer name
            first_num_idx = None
            for i, t in enumerate(tokens):
                if re.match(r"^[\d.,()-]+$", t):
                    first_num_idx = i
                    break

            if first_num_idx is None or first_num_idx == 0:
                continue

            ins_name = " ".join(tokens[:first_num_idx])
            if not ins_name:
                continue
            try:
                float(ins_name.replace(",", "").replace("-", "0"))
                continue
            except ValueError:
                pass

            canonical = normalize_insurer(ins_name)
            if canonical == "Unknown":
                continue
            if canonical not in result:
                result[canonical] = {}

            # Values after insurer name
            vals = tokens[first_num_idx:]
            # If duplicates: ICR values are the LAST n_icr_cols
            # If no duplicates: all values are ICR
            if has_duplicates:
                if len(vals) >= n_icr_cols:
                    icr_vals = vals[-n_icr_cols:]
                    for i, fy in enumerate(icr_years):
                        if i < len(icr_vals):
                            val = parse_num(icr_vals[i])
                            if val is not None:
                                result[canonical][fy] = val
            else:
                for i, fy in enumerate(icr_years):
                    if i < len(vals):
                        val = parse_num(vals[i])
                        if val is not None:
                            result[canonical][fy] = val

    return result


# ══════════════════════════════════════════════════════════════════════════════
#  SOURCE DISCOVERY
# ══════════════════════════════════════════════════════════════════════════════


def list_handbooks() -> list[dict]:
    handbooks = []
    if not HANDBOOK_DIR.exists():
        log.warning(f"Handbook directory not found: {HANDBOOK_DIR}")
        return handbooks

    for entry in sorted(HANDBOOK_DIR.iterdir()):
        name = entry.name

        # Match year from filename
        m = re.search(r"(\d{4})-(\d{2})", name)
        if not m:
            continue
        start_yr = int(m.group(1))
        end_yr = int(f"20{m.group(2)}")

        if entry.is_dir():
            continue

        if entry.suffix == ".pdf":
            handbooks.append(
                {"year": end_yr, "format": "pdf", "path": entry, "name": name}
            )
        elif entry.suffix == ".zip":
            handbooks.append(
                {"year": end_yr, "format": "zip", "path": entry, "name": name}
            )

    handbooks.sort(key=lambda h: h["year"])
    return handbooks


def identify_xlsx_structure(zip_path: Path) -> Optional[dict]:
    """Identify internal paths of Part II and Part III inside a handbook zip."""
    try:
        with zipfile.ZipFile(zip_path) as z:
            names = z.namelist()
    except Exception:
        return None

    part_ii = None
    part_iii = None
    summary = None
    has_xlsx = False
    has_xls = False

    for n in names:
        n_lower = n.lower()
        if "hindi" in n_lower:
            continue
        if n_lower.endswith(".xlsx"):
            has_xlsx = True
        if n_lower.endswith(".xls"):
            has_xls = True
        base = os.path.basename(n_lower)

        if n_lower.endswith(".xlsx"):
            part2_pattern = r"(?:(?<=^)|(?<=[\s_/\\]))part[\s_]+(?:ii(?!i)|2(?:nd)?)"
            part3_pattern = r"(?:(?<=^)|(?<=[\s_/\\]))part[\s_]+(?:iii|3(?:rd)?)"
            if re.search(part2_pattern, base):
                part_ii = n
            elif re.search(part3_pattern, base):
                part_iii = n
            elif "summary" in base and "index" not in base:
                summary = n

    if not part_ii and has_xlsx:
        for n in names:
            if "hindi" in n.lower():
                continue
            base = os.path.basename(n.lower())
            if "non-life" in base or "non_life" in base:
                part_ii = n
                break

    return {
        "part_ii": part_ii,
        "part_iii": part_iii,
        "summary": summary,
        "has_xlsx": has_xlsx,
        "has_xls": has_xls,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  XLSX EXTRACTOR (2020-21 to 2024-25)
# ══════════════════════════════════════════════════════════════════════════════


def extract_xlsx_gwp(wb, sheet_name: str, handbook_year: int) -> dict:
    """Extract Gross Written Premium from GWP table."""
    ws = wb[sheet_name]
    result = {}

    insurer_col = None
    year_cols = {}
    headers_found = False

    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)
        ]
        for ci, val in enumerate(row_vals):
            if insurer_col is None and val.lower() in (
                "insurer",
                "insurers",
                "general insurers",
            ):
                insurer_col = ci
            if re.match(r"^\d{4}-\d{2}$", val):
                year_cols[ci] = val

    if insurer_col is None:
        log.warning(f"  XLSX GWP: could not find insurer column in sheet {sheet_name}")
        return result

    for r in range(1, ws.max_row + 1):
        ins_name = str(ws.cell(r, insurer_col + 1).value or "").strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for ci, fy in year_cols.items():
            val = parse_num(ws.cell(r, ci + 1).value)
            if val is not None:
                result[canonical][fy] = val

    return result


def extract_xlsx_icr(
    wb, sheet_name: str, handbook_year: int, health_only: bool = False
) -> dict:
    """Extract Incurred Claims Ratio from ICR table."""
    ws = wb[sheet_name]
    result = {}

    insurer_col = None
    fy_icr_cols = {}

    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)
        ]
        for ci, val in enumerate(row_vals):
            if insurer_col is None and val.lower() in ("insurer", "insurers"):
                insurer_col = ci
            if re.match(r"^\d{4}-\d{2}$", val):
                fy_icr_cols[ci] = val

    # If no year headers found, try to find them by looking for "ICR" columns
    if not fy_icr_cols:
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [
                str(ws.cell(r, c).value or "").strip()
                for c in range(1, ws.max_column + 1)
            ]
            found_icr = False
            for ci, val in enumerate(row_vals):
                if "incurred claims ratio" in val.lower() or "icr" in val.lower():
                    found_icr = True
                if found_icr and re.match(r"^\d{4}-\d{2}$", val):
                    fy_icr_cols[ci] = val

    if insurer_col is None or not fy_icr_cols:
        log.warning(f"  XLSX ICR: could not locate columns in sheet {sheet_name}")
        return result

    for r in range(1, ws.max_row + 1):
        ins_name = str(ws.cell(r, insurer_col + 1).value or "").strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for ci, fy in fy_icr_cols.items():
            val = parse_num(ws.cell(r, ci + 1).value)
            if val is not None:
                result[canonical][fy] = val

    return result


def extract_xlsx_solvency(wb, sheet_name: str, handbook_year: int) -> dict:
    """Extract solvency ratio — use March-end (end of FY) values."""
    ws = wb[sheet_name]
    result = {}

    insurer_col = None
    march_cols = {}

    for r in range(1, min(5, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)
        ]
        for ci, val in enumerate(row_vals):
            if insurer_col is None and val.lower() in ("insurer", "insurers"):
                insurer_col = ci
            if "march" in val.lower() and re.search(r"\d{4}", val):
                m = re.search(r"(\d{4})", val)
                if m:
                    yr = int(m.group(1))
                    fy_key = f"{yr - 1}-{str(yr)[2:]}"
                    march_cols[ci] = fy_key

    if not march_cols:
        for r in range(1, min(5, ws.max_row + 1)):
            row_vals = [
                str(ws.cell(r, c).value or "").strip()
                for c in range(1, ws.max_column + 1)
            ]
            for ci, val in enumerate(row_vals):
                if val.lower() in (
                    "march 2015",
                    "march 2016",
                    "march 2017",
                    "march 2018",
                    "march 2019",
                    "march 2020",
                    "march 2021",
                    "march 2022",
                    "march 2023",
                    "march 2024",
                    "march 2025",
                ):
                    m = re.search(r"(\d{4})", val)
                    if m:
                        yr = int(m.group(1))
                        fy_key = f"{yr - 1}-{str(yr)[2:]}"
                        march_cols[ci] = fy_key

    if insurer_col is None:
        log.warning(f"  XLSX Solvency: no insurer column")
        return result

    for r in range(1, ws.max_row + 1):
        ins_name = str(ws.cell(r, insurer_col + 1).value or "").strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for ci, fy in march_cols.items():
            val = parse_num(ws.cell(r, ci + 1).value)
            if val is not None:
                result[canonical][fy] = val

    return result


def extract_xlsx_claims(wb, sheet_name: str, handbook_year: int) -> dict:
    """Extract claims data for CSR computation: claims paid, claims intimated.
    Handles multi-row headers where years are in one row and column descriptions in next row(s).
    """
    ws = wb[sheet_name]
    result = {}

    insurer_col = None
    claims_intimated_cols = {}
    claims_paid_cols = {}

    # Step 1: Find the year header row(s)
    year_row_map = {}
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)
        ]
        for ci, val in enumerate(row_vals):
            if insurer_col is None and val.lower() in ("insurer", "insurers"):
                insurer_col = ci
            m = re.match(r"^(\d{4}-\d{2})$", val)
            if m:
                year_row_map.setdefault(r, {})[ci] = m.group(1)

    if not year_row_map:
        log.warning(f"  XLSX Claims: no year headers found")
        return result

    # Determine which row has the most year headers (the main year header row)
    main_year_row = max(year_row_map, key=lambda k: len(year_row_map[k]))
    year_positions = year_row_map[main_year_row]

    # Step 2: Scan ALL header rows for "intimated", "booked", "paid" column labels
    # and associate them with the nearest year header to the left
    header_rows = list(range(1, min(15, ws.max_row + 1)))
    fy_ranges = sorted(year_positions.keys())

    for r in header_rows:
        row_vals = [
            str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)
        ]
        for ci, val in enumerate(row_vals):
            val_lower = val.lower()
            # Find which FY this column belongs to (nearest FY header above or to the left)
            fy = None
            for fy_col in sorted(year_positions.keys()):
                if fy_col <= ci:
                    fy = year_positions[fy_col]
                else:
                    break

            if fy and ("intimated" in val_lower or "booked" in val_lower):
                if (
                    "claims intimated" in val_lower
                    or "claims booked" in val_lower
                    or "intimated" in val_lower
                ):
                    claims_intimated_cols[ci] = fy
            if (
                fy
                and "paid" in val_lower
                and "re" not in val_lower
                and "repudiated" not in val_lower
            ):
                claims_paid_cols[ci] = fy

    if insurer_col is None:
        log.warning(f"  XLSX Claims: no insurer column")
        return result
    if not claims_intimated_cols or not claims_paid_cols:
        # Fallback: try adjacent column detection
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [
                str(ws.cell(r, c).value or "").strip()
                for c in range(1, ws.max_column + 1)
            ]
            for ci, val in enumerate(row_vals):
                val_lower = val.lower()
                if "claims intimated" in val_lower or "claims booked" in val_lower:
                    for c2 in range(ci - 5, ci):
                        if c2 >= 0 and re.match(r"^\d{4}-\d{2}$", row_vals[c2]):
                            claims_intimated_cols[ci] = row_vals[c2]
                if (
                    "claims paid" in val_lower
                    and "re" not in val_lower
                    and "repudiated" not in val_lower
                ):
                    for c2 in range(ci - 5, ci):
                        if c2 >= 0 and re.match(r"^\d{4}-\d{2}$", row_vals[c2]):
                            claims_paid_cols[ci] = row_vals[c2]

    if not claims_intimated_cols or not claims_paid_cols:
        log.warning(
            f"  XLSX Claims: no claims columns found (intimated={bool(claims_intimated_cols)}, paid={bool(claims_paid_cols)})"
        )
        return result

    for r in range(1, ws.max_row + 1):
        ins_name = str(ws.cell(r, insurer_col + 1).value or "").strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for ci, fy in claims_intimated_cols.items():
            val = parse_num(ws.cell(r, ci + 1).value)
            if val is not None:
                if fy not in result[canonical]:
                    result[canonical][fy] = {}
                result[canonical][fy]["claims_intimated"] = val
        for ci, fy in claims_paid_cols.items():
            val = parse_num(ws.cell(r, ci + 1).value)
            if val is not None:
                if fy not in result[canonical]:
                    result[canonical][fy] = {}
                result[canonical][fy]["claims_paid"] = val

    return result


def extract_xlsx_grievances(wb, sheet_name: str, handbook_year: int) -> dict:
    """Extract complaint counts from grievances table.
    Handles multi-row headers: years in one row, column descriptors in next row(s).
    """
    ws = wb[sheet_name]
    result = {}

    insurer_col = None
    reported_cols = {}

    # Step 1: Find year header row(s) and insurer column
    year_row_map = {}
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)
        ]
        for ci, val in enumerate(row_vals):
            if insurer_col is None and val.lower() in ("insurer", "insurers"):
                insurer_col = ci
            m = re.match(r"^(\d{4}-\d{2})$", val)
            if m:
                year_row_map.setdefault(r, {})[ci] = m.group(1)

    if not year_row_map:
        log.warning(f"  XLSX Grievances: no year headers found")
        return result

    main_year_row = max(year_row_map, key=lambda k: len(year_row_map[k]))
    year_positions = year_row_map[main_year_row]

    # Step 2: Scan all header rows for "Reported during the year" / "Actual Complaints"
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)
        ]
        for ci, val in enumerate(row_vals):
            val_lower = val.lower()
            if "reported" in val_lower or "actual complaints" in val_lower:
                fy = None
                for fy_col in sorted(year_positions.keys()):
                    if fy_col <= ci:
                        fy = year_positions[fy_col]
                if fy and ci not in reported_cols:
                    reported_cols[ci] = fy

    if insurer_col is None:
        log.warning(f"  XLSX Grievances: no insurer column")
        return result

    for r in range(1, ws.max_row + 1):
        ins_name = str(ws.cell(r, insurer_col + 1).value or "").strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for ci, fy in reported_cols.items():
            val = parse_num(ws.cell(r, ci + 1).value)
            if val is not None:
                result[canonical][fy] = val

    return result


def extract_xlsx_health_persons(wb, sheet_name: str, handbook_year: int) -> dict:
    """Extract lives covered (persons covered) from health table.
    Handles multi-row headers: years in one row, column descriptors in subsequent row(s).
    """
    ws = wb[sheet_name]
    result = {}

    insurer_col = None
    persons_cols = {}

    # Step 1: Find year header row(s)
    year_row_map = {}
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(r, c).value or "").strip()
            for c in range(1, min(30, ws.max_column + 1))
        ]
        for ci, val in enumerate(row_vals):
            if insurer_col is None and val.lower() in ("insurer", "insurers"):
                insurer_col = ci
            m = re.match(r"^(\d{4}-\d{2})$", val)
            if m:
                year_row_map.setdefault(r, {})[ci] = m.group(1)

    if not year_row_map:
        log.warning(f"  XLSX Health Persons: no year headers found")
        return result

    main_year_row = max(year_row_map, key=lambda k: len(year_row_map[k]))
    year_positions = year_row_map[main_year_row]

    # Step 2: Scan all header rows for "persons covered" descriptors
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(r, c).value or "").strip()
            for c in range(1, min(30, ws.max_column + 1))
        ]
        for ci, val in enumerate(row_vals):
            val_lower = val.lower()
            if "persons covered" in val_lower or "lives covered" in val_lower:
                fy = None
                for fy_col in sorted(year_positions.keys()):
                    if fy_col <= ci:
                        fy = year_positions[fy_col]
                if fy:
                    persons_cols[ci] = fy

    # Fallback: broader match
    if not persons_cols:
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [
                str(ws.cell(r, c).value or "").strip()
                for c in range(1, min(30, ws.max_column + 1))
            ]
            for ci, val in enumerate(row_vals):
                val_lower = val.lower()
                if "persons" in val_lower and (
                    "covered" in val_lower or "('000s)" in val_lower
                ):
                    for fy_col in sorted(year_positions.keys()):
                        if fy_col <= ci:
                            persons_cols[ci] = year_positions[fy_col]
                        else:
                            break

    if insurer_col is None:
        log.warning(f"  XLSX Health Persons: no insurer column")
        return result

    for r in range(1, ws.max_row + 1):
        ins_name = str(ws.cell(r, insurer_col + 1).value or "").strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for ci, fy in persons_cols.items():
            val = parse_num(ws.cell(r, ci + 1).value)
            if val is not None:
                result[canonical][fy] = val

    return result


def extract_xlsx_tpa_network(wb, sheet_name: str, handbook_year: int) -> dict:
    """Extract TPA-wise network hospital data."""
    ws = wb[sheet_name]
    result = {}

    tpa_col = None
    hospital_cols = {}

    for r in range(1, min(5, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)
        ]
        for ci, val in enumerate(row_vals):
            if tpa_col is None and ("tpa" in val.lower() or "name of" in val.lower()):
                tpa_col = ci
            if "number of hospitals" in val.lower() or "network" in val.lower():
                m = re.search(r"(\d{4})", val)
                if m:
                    hospital_cols[ci] = f"hospitals_{m.group(1)}"

    if tpa_col is None:
        log.warning(f"  XLSX TPA: no TPA name column")
        return result

    for r in range(1, ws.max_row + 1):
        tpa_name = str(ws.cell(r, tpa_col + 1).value or "").strip()
        if not tpa_name or "total" in tpa_name.lower():
            continue
        if tpa_name not in result:
            result[tpa_name] = {}
        for ci, key in hospital_cols.items():
            val = parse_num(ws.cell(r, ci + 1).value)
            if val is not None:
                result[tpa_name][key] = val

    return result


def process_xlsx_handbook(zip_path: Path, handbook_year: int) -> dict:
    """Process an XLSX handbook zip and return all extracted data."""
    data = {
        "gwp": {},
        "icr": {},
        "icr_health": {},
        "solvency": {},
        "claims": {},
        "grievances": {},
        "health_persons": {},
        "tpa_network": {},
    }

    structure = identify_xlsx_structure(zip_path)
    if not structure:
        log.warning(f"  Could not identify XLSX structure for {handbook_year}")
        return data

    tables = XLSX_PART_II_TABLES.get(handbook_year, XLSX_PART_II_TABLES[2025])
    health_tables = XLSX_PART_III_TABLES.get(handbook_year, XLSX_PART_III_TABLES[2025])

    tmpdir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(zip_path) as z:
            if structure["part_ii"]:
                z.extract(structure["part_ii"], tmpdir)
                extract_path = Path(tmpdir) / structure["part_ii"]
                wb = openpyxl.load_workbook(extract_path, data_only=True)

                if tables["gwp"] and tables["gwp"] in wb.sheetnames:
                    data["gwp"] = extract_xlsx_gwp(wb, tables["gwp"], handbook_year)
                if tables["icr_all"] and tables["icr_all"] in wb.sheetnames:
                    data["icr"] = extract_xlsx_icr(wb, tables["icr_all"], handbook_year)
                if tables["solvency"] and tables["solvency"] in wb.sheetnames:
                    data["solvency"] = extract_xlsx_solvency(
                        wb, tables["solvency"], handbook_year
                    )
                if tables["claims"] and tables["claims"] in wb.sheetnames:
                    data["claims"] = extract_xlsx_claims(
                        wb, tables["claims"], handbook_year
                    )
                if tables["grievances"] and tables["grievances"] in wb.sheetnames:
                    data["grievances"] = extract_xlsx_grievances(
                        wb, tables["grievances"], handbook_year
                    )
                if (
                    tables.get("health_persons")
                    and tables["health_persons"] in wb.sheetnames
                ):
                    data["health_persons"] = extract_xlsx_health_persons(
                        wb, tables["health_persons"], handbook_year
                    )
                if tables.get("health_icr") and tables["health_icr"] in wb.sheetnames:
                    data["icr_health"] = extract_xlsx_icr(
                        wb, tables["health_icr"], handbook_year
                    )

                wb.close()

            if structure["part_iii"]:
                z.extract(structure["part_iii"], tmpdir)
                extract_path = Path(tmpdir) / structure["part_iii"]
                wb = openpyxl.load_workbook(extract_path, data_only=True)

                if (
                    health_tables["health_persons"]
                    and health_tables["health_persons"] in wb.sheetnames
                ):
                    data["health_persons"] = extract_xlsx_health_persons(
                        wb, health_tables["health_persons"], handbook_year
                    )
                if (
                    health_tables["health_icr"]
                    and health_tables["health_icr"] in wb.sheetnames
                ):
                    data["icr_health"] = extract_xlsx_icr(
                        wb, health_tables["health_icr"], handbook_year
                    )
                if (
                    health_tables["tpa_network"]
                    and health_tables["tpa_network"] in wb.sheetnames
                ):
                    data["tpa_network"] = extract_xlsx_tpa_network(
                        wb, health_tables["tpa_network"], handbook_year
                    )

                wb.close()

    except Exception as e:
        log.warning(f"  XLSX processing error for {handbook_year}: {e}")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    return data


# ══════════════════════════════════════════════════════════════════════════════
#  XLS EXTRACTOR (2009-10, 2017-18)
# ══════════════════════════════════════════════════════════════════════════════


def extract_xls_simple(
    ws, name_col: int = 0, data_start_row: int = 1, year_col_map: Optional[dict] = None
) -> dict:
    """Generic XLS table parser for simple insurer → yearly-values tables."""
    result = {}
    for r in range(data_start_row, ws.nrows):
        ins_name = str(ws.cell_value(r, name_col)).strip()
        if not ins_name or is_total_row(ins_name):
            continue
        is_pct = False
        if ins_name.startswith("(") and ins_name.endswith(")"):
            continue
        try:
            float(ins_name.replace("-", "0"))
            continue
        except ValueError:
            pass
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        if year_col_map:
            for ci, fy in year_col_map.items():
                val = parse_num(ws.cell_value(r, ci))
                if val is not None:
                    result[canonical][fy] = val
    return result


def extract_xls_gwp(ws) -> dict:
    year_cols = {}
    for c in range(ws.ncols):
        val = str(ws.cell_value(3, c)).strip()
        m = re.match(r"^(\d{4})-(\d{2})$", val)
        if not m:
            m = re.match(r"^(\d{4})-(\d{4})$", val)
            if m:
                val = f"{m.group(1)}-{m.group(2)[2:]}"
            else:
                m = re.match(r"^(\d{4})$", str(val))
                if m:
                    yr = int(m.group(1))
                    val = f"{yr - 1}-{str(yr)[2:]}"
        if re.match(r"^\d{4}-\d{2}$", val) if not m else True:
            if re.match(r"^\d{4}-\d{2}$", str(ws.cell_value(3, c)).strip()):
                year_cols[c] = str(ws.cell_value(3, c)).strip()
            elif re.match(r"^\d{4}$", str(ws.cell_value(3, c)).strip()):
                yr = int(ws.cell_value(3, c))
                year_cols[c] = f"{yr - 1}-{str(yr)[2:]}"

    if not year_cols:
        for c in range(ws.ncols):
            val = str(ws.cell_value(3, c)).strip()
            m = re.match(r"^(\d{4})$", val)
            if m:
                yr = int(m.group(1))
                year_cols[c] = f"{yr - 1}-{str(yr)[2:]}"

    return extract_xls_simple(ws, name_col=0, data_start_row=4, year_col_map=year_cols)


def extract_xls_icr(ws) -> dict:
    """ICR in XLS has NEP and Claims Incurred side by side. Extract ICR values."""
    result = {}
    fy_columns = {}

    for c in range(ws.ncols):
        val = str(ws.cell_value(4, c)).strip() if ws.nrows > 4 else ""
        m = re.match(r"^(\d{4})-(\d{2})$", val)
        if not m:
            m = re.match(r"^(\d{4})$", val)
            if m:
                yr = int(m.group(1))
                val_fy = f"{yr - 1}-{str(yr)[2:]}"
                if val_fy not in fy_columns.values():
                    fy_columns[c] = val_fy

    if not fy_columns:
        return result

    current_section = ""
    for r in range(5, ws.nrows):
        ins_name = str(ws.cell_value(r, 0)).strip().upper()
        if not ins_name:
            continue
        if ins_name in ("PUBLIC", "PRIVATE", "SUB-TOTAL", "TOTAL"):
            current_section = ins_name
            continue

        # Check if this is an ICR row (look for % or ratio values)
        has_icr = False
        for c in fy_columns:
            val = str(ws.cell_value(r, c)).strip()
            if val and val != "0.0" and val != "0":
                has_icr = True
                break

        if not has_icr:
            continue

        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for c, fy in fy_columns.items():
            val = parse_num(ws.cell_value(r, c))
            if val is not None:
                result[canonical][fy] = val

    return result


def extract_xls_grievances_2010(ws) -> dict:
    result = {}
    year_cols = {}
    for c in range(ws.ncols):
        val = str(ws.cell_value(2, c)).strip() if ws.nrows > 2 else ""
        m = re.match(r"^(\d{4})$", val)
        if m:
            yr = int(m.group(1))
            year_cols[c] = f"{yr - 1}-{str(yr)[2:]}"
    for r in range(3, ws.nrows):
        ins_name = str(ws.cell_value(r, 0)).strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for c, fy in year_cols.items():
            val = parse_num(ws.cell_value(r, c))
            if val is not None:
                result[canonical][fy] = val
    return result


def extract_xls_grievances(ws) -> dict:
    """XLS grievances - find 'Reported during the year' or 'Actual Complaints' columns."""
    result = {}
    insurer_col = None
    reported_cols = {}

    for r in range(min(8, ws.nrows)):
        current_fy = None
        for c in range(ws.ncols):
            val = str(ws.cell_value(r, c)).strip()
            if insurer_col is None and val.lower() in ("insurer", "insurers"):
                insurer_col = c
            m = re.match(r"^(\d{4})-(\d{2})$", val)
            if m:
                current_fy = m.group(1)
            if current_fy and (
                "reported" in val.lower() or "actual complaints" in val.lower()
            ):
                if c not in reported_cols:
                    reported_cols[c] = current_fy

    if not reported_cols:
        for r in range(min(8, ws.nrows)):
            for c in range(ws.ncols):
                val = str(ws.cell_value(r, c)).strip()
                if "reported" in val.lower():
                    for c2 in range(c - 5, c):
                        if c2 >= 0:
                            m = re.match(
                                r"^(\d{4}-\d{2})$", str(ws.cell_value(r, c2)).strip()
                            )
                            if m:
                                reported_cols[c] = m.group(1)
                                break

    if insurer_col is None:
        insurer_col = 1

    for r in range(1, ws.nrows):
        ins_name = str(ws.cell_value(r, insurer_col)).strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for c, fy in reported_cols.items():
            if c < ws.ncols:
                val = parse_num(ws.cell_value(r, c))
                if val is not None:
                    result[canonical][fy] = val

    return result


def extract_xls_health_persons(ws) -> dict:
    result = {}
    fy_cols = {}
    insurer_col = 0

    for r in range(min(8, ws.nrows)):
        for c in range(ws.ncols):
            val = str(ws.cell_value(r, c)).strip()
            m = re.match(r"^(\d{4}-\d{2})$", val)
            if m:
                fy_cols[c] = m.group(1)
                break
        if fy_cols:
            break

    for r in range(3, ws.nrows):
        ins_name = str(ws.cell_value(r, insurer_col)).strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown" or canonical in PSU_INSURERS:
            continue
        if canonical not in result:
            result[canonical] = {}
        for c, fy in fy_cols.items():
            val = parse_num(ws.cell_value(r, c))
            if val is not None:
                result[canonical][fy] = val

    return result


def _read_spreadsheet(path: Path):
    """Try to read a spreadsheet file; returns (workbook, is_xlsx) or raises.
    Handles .xls files that are actually xlsx format internally.
    """
    # Determine actual format from magic bytes
    is_actually_xlsx = False
    try:
        with open(path, "rb") as f:
            header = f.read(8)
        if header[:2] == b"PK":
            is_actually_xlsx = True
    except Exception:
        pass

    if is_actually_xlsx:
        renamed = path.with_suffix(".xlsx")
        import shutil as _su

        try:
            _su.copy2(str(path), str(renamed))
            wb = openpyxl.load_workbook(renamed, data_only=True)
            return wb, True
        except Exception:
            pass
        finally:
            if renamed.exists():
                renamed.unlink()

    try:
        wb = xlrd.open_workbook(str(path))
        return wb, False
    except Exception:
        pass
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        return wb, True
    except Exception:
        pass
    raise ValueError(f"Cannot read {path}")


def _compute_icr_from_nep_claims(wb, sheet_name: str) -> dict:
    """For 2017-18 ICR TOTAL sheet: extract paired NEP+Claims columns and compute ICR%."""
    ws = wb[sheet_name]
    result = {}

    # Find year columns in row 4 (1-indexed). The sheet has:
    # Col A: INSURER
    # Col B-F: NET PREMIUM EARNED (2013-14..2017-18)
    # Col G-K: CLAIMS INCURRED (NET) (2013-14..2017-18)
    year_cols = []
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(4, c).value or "").strip()
        if re.match(r"^\d{4}-\d{2}$", val):
            year_cols.append((c, val))

    if not year_cols:
        return result

    n = len(year_cols) // 2
    if n == 0:
        return result

    nep_cols = year_cols[:n]
    claims_cols = year_cols[n:]

    for r in range(5, ws.max_row + 1):
        ins_name = str(ws.cell(r, 1).value or "").strip()
        if not ins_name or is_total_row(ins_name):
            continue
        canonical = normalize_insurer(ins_name)
        if canonical == "Unknown":
            continue
        if canonical not in result:
            result[canonical] = {}
        for i in range(n):
            nep_val = parse_num(ws.cell(r, nep_cols[i][0]).value)
            claims_val = parse_num(ws.cell(r, claims_cols[i][0]).value)
            fy = nep_cols[i][1]
            if nep_val is not None and nep_val != 0 and claims_val is not None:
                icr = (claims_val / nep_val) * 100
                result[canonical][fy] = round(icr, 2)

    return result


def process_xls_handbook(zip_path: Path, handbook_year: int) -> dict:
    data = {
        "gwp": {},
        "icr": {},
        "icr_health": {},
        "solvency": {},
        "claims": {},
        "grievances": {},
        "health_persons": {},
        "tpa_network": {},
    }

    tables_config = XLS_2018_TABLES if handbook_year >= 2018 else XLS_2010_TABLES
    is_2010 = handbook_year < 2018

    try:
        with zipfile.ZipFile(zip_path) as z:
            # Collect ICR file candidates, preferring TOTAL for 2017-18
            icr_candidates = []
            for fname_candidate in z.namelist():
                fbase_candidate = os.path.basename(fname_candidate).lower()
                icr_kw = tables_config.get("icr", "").lower()
                if icr_kw and icr_kw in fbase_candidate:
                    icr_candidates.append(fname_candidate)
            # Prefer "total" over "health"/"pa" variants for 2017-18
            if handbook_year >= 2018:
                for c in icr_candidates:
                    if "total" in os.path.basename(c).lower():
                        icr_candidates.remove(c)
                        icr_candidates.insert(0, c)
                        break

            for fname in z.namelist():
                fbase = os.path.basename(fname)
                fname_lower = fbase.lower()

                if tables_config["gwp"].lower() in fname_lower:
                    tmpdir = tempfile.mkdtemp()
                    try:
                        z.extract(fname, tmpdir)
                        wb_obj, is_xlsx = _read_spreadsheet(Path(tmpdir) / fname)
                        ws = (
                            wb_obj.sheet_by_index(0)
                            if not is_xlsx
                            else wb_obj[wb_obj.sheetnames[0]]
                        )
                        if is_xlsx:
                            data["gwp"] = extract_xlsx_gwp(
                                wb_obj, wb_obj.sheetnames[0], handbook_year
                            )
                        else:
                            data["gwp"] = extract_xls_gwp(ws)
                    except Exception as e:
                        log.warning(f"  GWP error in {handbook_year}: {e}")
                    finally:
                        shutil.rmtree(tmpdir, ignore_errors=True)

                elif icr_candidates and fname == icr_candidates[0]:
                    tmpdir = tempfile.mkdtemp()
                    try:
                        z.extract(fname, tmpdir)
                        wb_obj, is_xlsx = _read_spreadsheet(Path(tmpdir) / fname)
                        if is_xlsx:
                            # Multi-sheet ICR (2017-18): find TOTAL sheet
                            sheet_to_use = wb_obj.sheetnames[0]
                            for sn in wb_obj.sheetnames:
                                if "total" in sn.lower():
                                    sheet_to_use = sn
                                    break
                            data["icr"] = extract_xlsx_icr(
                                wb_obj, sheet_to_use, handbook_year
                            )
                            # 2017-18 TOTAL sheet has NEP+Claims columns, not ICR%.
                            # extract_xlsx_icr gets Claims values (overwrites NEP).
                            # Compute ICR = claims / NEP * 100 from paired columns.
                            if (
                                handbook_year == 2018
                                and "total" in sheet_to_use.lower()
                            ):
                                computed = _compute_icr_from_nep_claims(
                                    wb_obj, sheet_to_use
                                )
                                if computed:
                                    data["icr"] = computed
                        else:
                            ws = wb_obj.sheet_by_index(0)
                            data["icr"] = extract_xls_icr(ws)
                    except Exception as e:
                        log.warning(f"  ICR error in {handbook_year}: {e}")
                    finally:
                        shutil.rmtree(tmpdir, ignore_errors=True)

                elif tables_config["solvency"].lower() in fname_lower:
                    tmpdir = tempfile.mkdtemp()
                    try:
                        z.extract(fname, tmpdir)
                        wb_obj, is_xlsx = _read_spreadsheet(Path(tmpdir) / fname)
                        ws = (
                            wb_obj.sheet_by_index(0)
                            if not is_xlsx
                            else wb_obj[wb_obj.sheetnames[0]]
                        )
                        data["solvency"] = extract_xls_simple(
                            ws, name_col=0, data_start_row=2
                        )
                    except Exception:
                        pass
                    finally:
                        shutil.rmtree(tmpdir, ignore_errors=True)

                elif tables_config["grievances"].lower() in fname_lower:
                    tmpdir = tempfile.mkdtemp()
                    try:
                        z.extract(fname, tmpdir)
                        wb_obj, is_xlsx = _read_spreadsheet(Path(tmpdir) / fname)
                        ws = (
                            wb_obj.sheet_by_index(0)
                            if not is_xlsx
                            else wb_obj[wb_obj.sheetnames[0]]
                        )
                        if is_2010:
                            data["grievances"] = extract_xls_grievances_2010(ws)
                        else:
                            data["grievances"] = extract_xls_grievances(ws)
                    except Exception:
                        pass
                    finally:
                        shutil.rmtree(tmpdir, ignore_errors=True)

                elif (
                    not is_2010
                    and "health_persons" in tables_config
                    and "persons" in fname_lower
                ):
                    tmpdir = tempfile.mkdtemp()
                    try:
                        z.extract(fname, tmpdir)
                        wb_obj, is_xlsx = _read_spreadsheet(Path(tmpdir) / fname)
                        ws = (
                            wb_obj.sheet_by_index(0)
                            if not is_xlsx
                            else wb_obj[wb_obj.sheetnames[0]]
                        )
                        data["health_persons"] = extract_xls_health_persons(ws)
                    except Exception:
                        pass
                    finally:
                        shutil.rmtree(tmpdir, ignore_errors=True)

                elif (
                    not is_2010
                    and "health_icr" in tables_config
                    and "incurred_claims_ratio" in fname_lower
                    and "hi_" in fname_lower
                ):
                    tmpdir = tempfile.mkdtemp()
                    try:
                        z.extract(fname, tmpdir)
                        wb_obj, is_xlsx = _read_spreadsheet(Path(tmpdir) / fname)
                        ws = (
                            wb_obj.sheet_by_index(0)
                            if not is_xlsx
                            else wb_obj[wb_obj.sheetnames[0]]
                        )
                        data["icr_health"] = extract_xls_icr(ws)
                    except Exception:
                        pass
                    finally:
                        shutil.rmtree(tmpdir, ignore_errors=True)

    except Exception as e:
        log.warning(f"  XLS processing error for {handbook_year}: {e}")

    return data


# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXTRACTOR (all PDF-era handbooks)
# ══════════════════════════════════════════════════════════════════════════════


def pdf_find_table_pages(pdf, keywords: list) -> list[tuple[int, str]]:
    """Find pages in PDF that match table keywords. Handles reversed text pages.
    Filters out TOC/contents pages that mention tables but don't contain data."""
    matches = []
    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        first_line = text.strip().split("\n")[0] if text.strip() else ""
        matched = False
        for kw in keywords:
            if re.search(kw, text[:500], re.IGNORECASE):
                matched = True
                break
        if not matched and _is_reversed_text(text[:2000]):
            fixed = _rebuild_page_text(page)
            fixed_joined = " ".join(fixed.split())
            for kw in keywords:
                if re.search(kw, fixed_joined[:1000], re.IGNORECASE):
                    matched = True
                    break
        if matched:
            # Skip TOC/contents pages (they list table numbers but have no data)
            if re.search(r"table\s+no", text[:300], re.IGNORECASE) and re.search(
                r"page\s+no", text[:300], re.IGNORECASE
            ):
                continue
            matches.append((i, first_line[:100]))
    return matches


def pdf_parse_table_text(text: str, year_cols: list[str]) -> list[dict]:
    """Parse space-delimited table text into structured records."""
    records = []
    lines = text.strip().split("\n")

    for line in lines:
        line = line.strip()
        if not line:
            continue
        if is_total_row(line[:50]):
            continue
        # Skip lines that are just footnotes or legends
        if re.match(r"^[A-Z\s]{3,}\d{1,2}\.\s*$", line):
            continue
        if re.match(r"^[A-Z\s]{3,}$", line) and len(line) < 30:
            continue
        if re.match(r"^[\d\s.,()%-]+$", line):
            continue

        records.append({"raw": line, "year_cols": year_cols})

    return records


def process_pdf_handbook(pdf_path: Path, handbook_year: int) -> dict:
    """Process a PDF handbook and extract all available data."""
    data = {
        "gwp": {},
        "icr": {},
        "icr_health": {},
        "solvency": {},
        "claims": {},
        "grievances": {},
        "health_persons": {},
        "tpa_network": {},
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            log.info(f"  PDF has {len(pdf.pages)} pages, scanning for tables...")

            # Find table pages
            gwp_pages = pdf_find_table_pages(pdf, PDF_TABLE_KWARGS["gwp"])
            icr_pages = pdf_find_table_pages(pdf, PDF_TABLE_KWARGS["icr"])

            if gwp_pages:
                log.info(f"  Found GWP on pages: {[p[0] + 1 for p in gwp_pages[:5]]}")
                for pg_num, title in gwp_pages:
                    page = pdf.pages[pg_num]
                    parsed = pdf_extract_gwp_table(page, handbook_year)
                    data["gwp"].update(parsed)

            if icr_pages:
                log.info(f"  Found ICR on pages: {[p[0] + 1 for p in icr_pages[:5]]}")
                for pg_num, title in icr_pages:
                    page = pdf.pages[pg_num]
                    # Filter to only "TOTAL"/"ALL SEGMENTS" ICR pages
                    pg_text = page.extract_text() or ""
                    is_segment_total = False
                    total_regex = r"(?:table\s+\d+|incurred claims ratio).{0,80}(?:total|all\s+segments)"
                    if _is_reversed_text(pg_text[:2000]):
                        rebuilt = _rebuild_page_text(page)
                        rebuilt_joined = " ".join(rebuilt.split())
                        if re.search(
                            total_regex,
                            rebuilt_joined[:1000],
                            re.IGNORECASE,
                        ):
                            is_segment_total = True
                    else:
                        text_joined = " ".join(pg_text.split())
                        if re.search(
                            total_regex,
                            text_joined[:1000],
                            re.IGNORECASE,
                        ):
                            is_segment_total = True
                    if not is_segment_total:
                        continue
                    # Try both extraction methods and take the larger result
                    landscape = _extract_landscape_table(page, handbook_year)
                    portrait = pdf_extract_icr_table(page, handbook_year)
                    parsed = landscape if len(landscape) >= len(portrait) else portrait
                    if parsed:
                        log.info(
                            f"    Page {pg_num + 1}: landscape={len(landscape)}, portrait={len(portrait)}, using {'landscape' if len(landscape) >= len(portrait) else 'portrait'}"
                        )
                    data["icr"].update(parsed)

    except Exception as e:
        log.warning(f"  PDF processing error for {handbook_year}: {e}")

    return data


# ══════════════════════════════════════════════════════════════════════════════
#  CSR COMPUTER
# ══════════════════════════════════════════════════════════════════════════════


def compute_csr(claims_data: dict, handbooks_data: list) -> dict:
    """Compute Claim Settlement Ratio from raw claims data across handbooks."""
    csr_result = {}

    for ins_data in claims_data:
        for canonical, fy_data in ins_data.items():
            if canonical not in csr_result:
                csr_result[canonical] = {}
            for fy, vals in fy_data.items():
                if isinstance(vals, dict):
                    intim = vals.get("claims_intimated")
                    paid = vals.get("claims_paid")
                    if intim and paid and float(intim) > 0:
                        csr_result[canonical][fy] = round(
                            (float(paid) / float(intim)) * 100, 2
                        )

    return csr_result


# ══════════════════════════════════════════════════════════════════════════════
#  DATA MERGER
# ══════════════════════════════════════════════════════════════════════════════


def is_valid_insurer_name(name: str) -> bool:
    """Second-layer check: is this name likely a real insurer (not a footnote/header artifact)?"""
    if not name or name == "Unknown":
        return False
    # Short names that are known canonical names are always valid (check BEFORE header check)
    canonical_values = {v.lower() for v in CANONICAL_NAMES.values()}
    if name.lower() in canonical_values:
        return True
    if is_footnote_or_header(name):
        return False
    # Anti-patterns: clearly not an insurer name
    name_lower = name.lower()
    anti_patterns = [
        "ombudsmen",
        "grievance",
        "consumer",
        "redressal",
        "density",
        "penetration",
        "policies",
        "performance of",
        "status of",
        "branch office",
        "registered",
        "profit/loss",
        "profit and loss",
        "channel wise",
        "state wise",
        "segment wise",
        "life and non-life",
        "life & non-life",
        "no of companies",
        "number of companies",
        "no of new",
        "number of new",
        "$",  # dollar sign = financial metric
        "handbook",
        "profile",
        "business figures",
        # PDF artifacts
        "and specialised insurers",
        "and specialized insurers",
        "– repres",  # em-dash footnote
    ]
    for pat in anti_patterns:
        if pat in name_lower:
            return False
    # Must have at least 2 words, OR be a known all-caps abbreviation
    words = name.split()
    # Any word that is just a digit (like "United 2", "United 3")
    for w in words:
        if w.isdigit():
            return False

    known_abbrevs = {"ECGC", "AIC", "LIC", "GIC", "PSU", "PPI"}
    if len(words) < 2:
        if name.upper() == name and name not in known_abbrevs:
            return False
        if name.lower() not in canonical_values:
            return False
    # All-caps with ≤3 words and no insurer keyword is likely a PDF artifact
    if name.upper() == name and len(words) <= 3:
        insurer_kw = {
            "insurance",
            "general",
            "health",
            "ltd",
            "co",
            "limited",
            "company",
            "india",
            "indian",
        }
        if not any(w.lower() in insurer_kw for w in words):
            return False
    # Must have at least one meaningful word (not all stop-words)
    stop_words = {"the", "of", "and", "in", "for", "to", "a", "an", "ltd", "co", "inc"}
    meaningful = [w for w in words if w.lower() not in stop_words]
    if not meaningful:
        return False
    return True


def merge_data(data_sets: list[dict]) -> dict:
    """Merge multiple handbook extractions, preferring newer handbooks for overlapping years."""
    merged = {}

    for ds in data_sets:
        handbook_year = ds.get("handbook_year", 0)
        for field in [
            "gwp",
            "icr",
            "icr_health",
            "solvency",
            "claims",
            "grievances",
            "health_persons",
        ]:
            for raw_key, fy_data in ds.get(field, {}).items():
                canonical = normalize_insurer(raw_key)
                if canonical == "Unknown" or not is_valid_insurer_name(canonical):
                    continue
                if canonical not in merged:
                    merged[canonical] = {}
                for fy, value in fy_data.items():
                    if fy not in merged[canonical]:
                        merged[canonical][fy] = {}
                    # Prefer newer handbook data
                    current_source = merged[canonical][fy].get("_source_year", 0)
                    if handbook_year >= current_source:
                        merged[canonical][fy][field] = value
                        merged[canonical][fy]["_source_year"] = handbook_year

    return merged


# ══════════════════════════════════════════════════════════════════════════════
#  CONSISTENCY ANALYZER
# ══════════════════════════════════════════════════════════════════════════════


def compute_consistency(csr_data: dict) -> dict:
    result = {}
    for canonical, fy_data in csr_data.items():
        values = [v for v in fy_data.values() if v is not None]
        if len(values) < 2:
            result[canonical] = {
                "csr_values": values,
                "csr_std_dev": None,
                "consistency_label": "Insufficient Data",
                "years_available": len(values),
            }
            continue
        mean = sum(values) / len(values)
        variance = sum((v - mean) ** 2 for v in values) / len(values)
        std_dev = variance**0.5

        if std_dev <= 2.0:
            label = "Very Consistent"
        elif std_dev <= 5.0:
            label = "Mostly Consistent"
        else:
            label = "Inconsistent"

        result[canonical] = {
            "csr_values": values,
            "csr_mean": round(mean, 2),
            "csr_std_dev": round(std_dev, 2),
            "consistency_label": label,
            "years_available": len(values),
        }
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════


def build_output(
    data_sets: list[dict],
    csr_data: dict,
    consistency: dict,
    tpa_data: dict,
    handbooks_processed: list,
) -> dict:
    """Build the final output JSON structure."""
    merged = merge_data(data_sets)
    all_years = set()
    for canonical, fy_data in merged.items():
        all_years.update(fy_data.keys())
    all_years.discard("_source_year")
    all_years = sorted(all_years)

    # Build historical_data
    historical_data = {}
    canonical_list = sorted(set(list(merged.keys()) + list(csr_data.keys())))
    for canonical in canonical_list:
        historical_data[canonical] = {}
        for fy in all_years:
            entry = merged.get(canonical, {}).get(fy, {})
            csr_val = (
                csr_data.get(canonical, {}).get(fy)
                if isinstance(csr_data.get(canonical), dict)
                else None
            )
            historical_data[canonical][fy] = {
                "gwp_cr": entry.get("gwp"),
                "icr_pct": entry.get("icr"),
                "icr_health_pct": entry.get("icr_health"),
                "solvency_ratio": entry.get("solvency"),
                "lives_covered": entry.get("health_persons"),
                "complaints_reported": entry.get("grievances"),
                "csr_pct": csr_val,
            }

    return {
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "handbooks_processed": handbooks_processed,
            "years_covered": all_years,
            "total_insurers": len(canonical_list),
        },
        "historical_data": historical_data,
        "consistency_scores": {k: v for k, v in consistency.items()},
        "tpa_network_data": tpa_data,
    }


def main():
    log.info("=" * 60)
    log.info("IRDAI Historical Performance Scraper")
    log.info("=" * 60)

    handbooks = list_handbooks()
    log.info(f"Found {len(handbooks)} handbooks")

    if not handbooks:
        log.warning(f"No handbooks found in {HANDBOOK_DIR}")
        log.info("Falling back to Annual Reports...")
        # TODO: annual report extraction (less structured, lower priority)

    data_sets = []
    tpa_data = {}
    handbooks_processed = []
    claims_data_sources = []

    for hb in handbooks:
        year = hb["year"]
        fmt = hb["format"]
        log.info(f"\nProcessing {hb['name']} (FY ending {year}, {fmt})...")

        if fmt == "zip":
            structure = identify_xlsx_structure(hb["path"])
            if structure and structure["has_xlsx"]:
                ds = process_xlsx_handbook(hb["path"], year)
                ds["handbook_year"] = year
                data_sets.append(ds)
                handbooks_processed.append(
                    {"name": hb["name"], "year": year, "format": "xlsx"}
                )

                if ds["claims"]:
                    claims_data_sources.append(ds["claims"])
                if ds["tpa_network"]:
                    tpa_data.update(ds["tpa_network"])

                log.info(
                    f"  GWP: {len(ds['gwp'])} insurers | ICR: {len(ds['icr'])} | "
                    f"Solvency: {len(ds['solvency'])} | Claims: {len(ds['claims'])} | "
                    f"Grievances: {len(ds['grievances'])} | HealthPersons: {len(ds['health_persons'])}"
                )
            elif structure and structure["has_xls"]:
                ds = process_xls_handbook(hb["path"], year)
                ds["handbook_year"] = year
                data_sets.append(ds)
                handbooks_processed.append(
                    {"name": hb["name"], "year": year, "format": "xls"}
                )
                log.info(
                    f"  GWP: {len(ds['gwp'])} insurers | ICR: {len(ds['icr'])} | "
                    f"Grievances: {len(ds['grievances'])}"
                )

        elif fmt == "pdf":
            ds = process_pdf_handbook(hb["path"], year)
            ds["handbook_year"] = year
            data_sets.append(ds)
            handbooks_processed.append(
                {"name": hb["name"], "year": year, "format": "pdf"}
            )
            log.info(f"  GWP: {len(ds['gwp'])} insurers | ICR: {len(ds['icr'])}")

        else:
            log.warning(f"  Unknown format: {fmt}")

    # Compute CSR
    log.info("\n" + "=" * 60)
    log.info("Computing Claim Settlement Ratio...")
    csr_data = compute_csr(claims_data_sources, data_sets)
    log.info(f"  CSR computed for {len(csr_data)} insurers")

    # Compute consistency
    consistency = compute_consistency(csr_data)

    # Build output
    output = build_output(
        data_sets, csr_data, consistency, tpa_data, handbooks_processed
    )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(output, indent=2, ensure_ascii=False, default=str)
    )
    log.info(f"\nWritten to {OUTPUT_PATH}")

    if tpa_data:
        TPA_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        TPA_OUTPUT_PATH.write_text(
            json.dumps(tpa_data, indent=2, ensure_ascii=False, default=str)
        )
        log.info(f"TPA network data written to {TPA_OUTPUT_PATH}")

    # Summary
    log.info("\n" + "=" * 60)
    log.info("SUMMARY")
    log.info("=" * 60)
    log.info(f"Handbooks processed: {len(handbooks_processed)}")
    log.info(f"Insurers in merged data: {len(output['historical_data'])}")
    log.info(
        f"Years covered: {output['metadata']['years_covered'][:3]} ... {output['metadata']['years_covered'][-3:]}"
    )
    log.info(
        f"Total year-insurer datapoints: {sum(len(fy) for fy in output['historical_data'].values())}"
    )
    log.info(f"Insurers with CSR: {len(csr_data)}")

    for canonical in sorted(consistency.keys())[:10]:
        c = consistency[canonical]
        if c.get("consistency_label"):
            log.info(
                f"  {canonical:40s} | {c['consistency_label']:20s} | CSR std: {c.get('csr_std_dev', 'N/A')} | Years: {c['years_available']}"
            )


if __name__ == "__main__":
    main()
